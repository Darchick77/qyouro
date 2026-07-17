import sys
import json
import requests
import os
from datetime import datetime

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QStackedWidget,
    QVBoxLayout, QHBoxLayout, QFormLayout, QGridLayout,
    QLabel, QLineEdit, QPushButton, QTableWidget, QTableWidgetItem,
    QHeaderView, QComboBox, QSpinBox, QMessageBox,
    QFrame, QMenu, QGroupBox, QTextEdit, QFileDialog,
    QAbstractItemView, QSystemTrayIcon, QStyle, QDialogButtonBox,
    QDialog, QDateEdit, QStatusBar,
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QDate
from PyQt6.QtGui import QAction, QColor, QKeySequence

API_URL = os.environ.get("QYOURO_API", "https://qyouro-1.onrender.com")

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

STYLE = """
QMainWindow { background: #0f1117; }
QWidget { color: #e1e4e8; font-family: 'Segoe UI', sans-serif; }
QFrame#sidebar { background: #161b22; border-right: 1px solid #21262d; }
QLabel#logo { color: #58a6ff; font-size: 18px; font-weight: bold; padding: 8px 12px; }
QLabel#section { color: #8b949e; font-size: 11px; font-weight: bold; margin-top: 8px; padding: 4px 12px; }
QLabel#page-title { font-size: 22px; font-weight: bold; color: #e1e4e8; margin-bottom: 4px; }
QLineEdit, QSpinBox, QComboBox, QDateEdit {
    background: #0d1117; color: #c9d1d9; border: 1px solid #30363d;
    border-radius: 6px; padding: 6px 10px; font-size: 13px;
}
QLineEdit:focus, QSpinBox:focus, QComboBox:focus { border-color: #58a6ff; }
QPushButton {
    background: #21262d; color: #c9d1d9; border: 1px solid #30363d;
    border-radius: 6px; padding: 7px 16px; font-size: 13px; font-weight: 500;
}
QPushButton:hover { background: #30363d; }
QPushButton#primary { background: #238636; color: #fff; border-color: #2ea043; font-weight: bold; }
QPushButton#primary:hover { background: #2ea043; }
QPushButton#danger { background: #da3633; color: #fff; border-color: #f85149; }
QPushButton#danger:hover { background: #f85149; }
QPushButton#warning { background: #d29922; color: #000; border-color: #e3b341; }
QPushButton#sidebar_btn {
    background: transparent; color: #8b949e; border: none; border-radius: 8px;
    text-align: left; padding: 10px 16px; font-size: 14px;
}
QPushButton#sidebar_btn:hover { background: #21262d; color: #e1e4e8; }
QPushButton#sidebar_btn:checked { background: #1f6feb22; color: #58a6ff; border-left: 3px solid #58a6ff; }
QTableWidget {
    background: #0d1117; color: #c9d1d9; border: 1px solid #30363d;
    gridline-color: #21262d; font-size: 12px; alternate-background-color: #161b22;
    selection-background-color: #1f6feb44;
}
QHeaderView::section {
    background: #161b22; color: #e1e4e8; border: none; border-bottom: 2px solid #30363d;
    padding: 8px; font-weight: bold; font-size: 11px;
}
QScrollBar:vertical { background: #0d1117; width: 8px; }
QScrollBar::handle:vertical { background: #30363d; border-radius: 4px; min-height: 30px; }
QStatusBar { background: #161b22; color: #8b949e; border-top: 1px solid #21262d; font-size: 12px; }
QMenu { background: #161b22; color: #e1e4e8; border: 1px solid #30363d; padding: 4px; }
QMenu::item { padding: 6px 24px; border-radius: 4px; }
QMenu::item:selected { background: #1f6feb; }
QGroupBox { color: #e1e4e8; border: 1px solid #30363d; border-radius: 8px; margin-top: 12px; padding-top: 16px; font-weight: bold; }
QGroupBox::title { subcontrol-origin: margin; left: 12px; padding: 0 6px; }
QComboBox QAbstractItemView { background: #0d1117; color: #c9d1d9; border: 1px solid #30363d; selection-background-color: #1f6feb; }
"""


class ApiWorker(QThread):
    finished = pyqtSignal(dict, str)

    def __init__(self, method, path, data=None, tag=""):
        super().__init__()
        self._m = method
        self._p = path
        self._d = data
        self._tag = tag

    def run(self):
        try:
            kwargs = {"timeout": 10}
            if self._d:
                if self._m in ("post", "put", "patch"):
                    kwargs["json"] = self._d
                else:
                    kwargs["params"] = self._d
            r = requests.request(self._m, f"{API_URL}{self._p}", **kwargs)
            data = r.json() if r.headers.get("content-type", "").startswith("application/json") else {"ok": False}
            self.finished.emit(data, self._tag)
        except Exception as e:
            self.finished.emit({"ok": False, "error": str(e)}, self._tag)


class StatCard(QFrame):
    def __init__(self, title, color="#58a6ff"):
        super().__init__()
        self.setStyleSheet(f"QFrame {{ background: #161b22; border: 1px solid #21262d; border-radius: 10px; padding: 14px; }}")
        self.setMinimumSize(145, 90)
        l = QVBoxLayout(self)
        l.setSpacing(4)
        self._val = QLabel("—")
        self._val.setStyleSheet(f"color: {color}; font-size: 28px; font-weight: bold;")
        self._val.setAlignment(Qt.AlignmentFlag.AlignCenter)
        l.addWidget(self._val)
        t = QLabel(title)
        t.setStyleSheet("color: #8b949e; font-size: 11px;")
        t.setAlignment(Qt.AlignmentFlag.AlignCenter)
        l.addWidget(t)

    def set_value(self, v):
        self._val.setText(str(v))


class QyouroAdmin(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Qyouro Admin")
        self.resize(1240, 780)
        self.setMinimumSize(980, 600)
        self.setStyleSheet(STYLE)
        self._workers = []
        self._sort_col = -1
        self._sort_asc = True
        self._build_ui()
        self._setup_tray()
        self._setup_shortcuts()
        self.show()
        self._refresh_dashboard()

    def _api(self, method, path, data=None, callback=None, tag=""):
        w = ApiWorker(method, path, data, tag)
        if callback:
            w.finished.connect(lambda r, t: callback(r))
        w.finished.connect(lambda: self._workers.remove(w) if w in self._workers else None)
        self._workers.append(w)
        w.start()
        return w

    def _status(self, msg, timeout=3000):
        self.statusBar().showMessage(msg, timeout)

    def _build_ui(self):
        c = QWidget()
        self.setCentralWidget(c)
        root = QHBoxLayout(c)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # === SIDEBAR ===
        side = QFrame(objectName="sidebar")
        side.setFixedWidth(200)
        sl = QVBoxLayout(side)
        sl.setContentsMargins(8, 12, 8, 12)
        sl.setSpacing(2)

        logo = QLabel("Qyouro Admin", objectName="logo")
        sl.addWidget(logo)

        sl.addWidget(QLabel("ГЛАВНАЯ", objectName="section"))
        self._btn_dash = self._side_btn("Дашборд", 0, True)
        self._btn_keys = self._side_btn("Ключи", 1)
        sl.addWidget(self._btn_dash)
        sl.addWidget(self._btn_keys)

        sl.addWidget(QLabel("УПРАВЛЕНИЕ", objectName="section"))
        self._btn_emp = self._side_btn("Сотрудники", 2)
        self._btn_audit = self._side_btn("Журнал", 3)
        sl.addWidget(self._btn_emp)
        sl.addWidget(self._btn_audit)

        sl.addStretch()
        restart_btn = QPushButton("Перезапустить сервер")
        restart_btn.setStyleSheet("background: transparent; color: #8b949e; border: 1px solid #30363d; border-radius: 6px; padding: 5px; font-size: 10px;")
        restart_btn.clicked.connect(self._restart_backend)
        sl.addWidget(restart_btn)

        self._all_btns = [self._btn_dash, self._btn_keys, self._btn_emp, self._btn_audit]
        root.addWidget(side)

        # === CONTENT STACK ===
        self._stack = QStackedWidget()
        self._stack.addWidget(self._page_dashboard())
        self._stack.addWidget(self._page_keys())
        self._stack.addWidget(self._page_employees())
        self._stack.addWidget(self._page_audit())
        root.addWidget(self._stack)

        self.statusBar().showMessage("Готов")

    def _side_btn(self, text, idx, checked=False):
        b = QPushButton(text, objectName="sidebar_btn")
        b.setCheckable(True)
        b.setChecked(checked)
        b.clicked.connect(lambda: self._nav(idx))
        return b

    def _nav(self, idx):
        for i, b in enumerate(self._all_btns):
            b.setChecked(i == idx)
        self._stack.setCurrentIndex(idx)
        names = ["Дашборд", "Ключи", "Сотрудники", "Журнал"]
        self._status(names[idx])
        if idx == 0:
            self._refresh_dashboard()
        elif idx == 1:
            self._refresh_keys()
        elif idx == 2:
            self._refresh_employees()
        elif idx == 3:
            self._refresh_audit()

    def _setup_tray(self):
        self._tray = QSystemTrayIcon(self)
        self._tray.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_ComputerIcon))
        m = QMenu()
        m.addAction("Показать", self.show)
        m.addSeparator()
        m.addAction("Выход", QApplication.quit)
        self._tray.setContextMenu(m)
        self._tray.activated.connect(lambda r: self.showNormal() if r == QSystemTrayIcon.ActivationReason.DoubleClick else None)
        self._tray.show()

    def closeEvent(self, e):
        if self._tray.isVisible():
            self.hide()
            self._tray.showMessage("Qyouro Admin", "Свёрнуто в трей", QSystemTrayIcon.MessageIcon.Information, 2000)
            e.ignore()
        else:
            e.accept()

    def _setup_shortcuts(self):
        for i, key in enumerate(["Ctrl+1", "Ctrl+2", "Ctrl+3", "Ctrl+4"]):
            a = QAction(self, shortcut=QKeySequence(key), triggered=lambda v=i: self._nav(v))
            self.addAction(a)
        self.addAction(QAction(self, shortcut=QKeySequence("Ctrl+E"), triggered=self._export_excel))
        self.addAction(QAction(self, shortcut=QKeySequence("Ctrl+Q"), triggered=QApplication.quit))

    def _restart_backend(self):
        self._api("get", "/health", callback=lambda r: self._status("Сервер: " + r.get("service", "OK")))

    # ═══════════════════════════ DASHBOARD ═══════════════════════════

    def _page_dashboard(self):
        p = QWidget()
        l = QVBoxLayout(p)
        l.setContentsMargins(24, 20, 24, 20)
        l.setSpacing(16)

        l.addWidget(QLabel("Дашборд", objectName="page-title"))

        g = QGridLayout()
        g.setSpacing(12)
        self._cards = {
            "total": StatCard("Всего ключей"),
            "active": StatCard("Активных", "#3fb950"),
            "expired": StatCard("Истекло", "#f85149"),
            "revoked": StatCard("Отозвано", "#8b949e"),
            "activated": StatCard("Активировано", "#a371f7"),
            "employees": StatCard("Сотрудников", "#79c0ff"),
            "expiring": StatCard("Истекают (7 дн.)", "#d29922"),
        }
        for i, (k, card) in enumerate(self._cards.items()):
            g.addWidget(card, i // 4, i % 4)
        l.addLayout(g)
        l.addStretch()
        return p

    def _refresh_dashboard(self):
        def cb(r):
            if r.get("total") is not None:
                self._cards["total"].set_value(r["total"])
                self._cards["active"].set_value(r["active"])
                self._cards["expired"].set_value(r["expired"])
                self._cards["revoked"].set_value(r["revoked"])
                self._cards["activated"].set_value(r["activated"])
                self._cards["employees"].set_value(r["employees"])
                self._cards["expiring"].set_value(r["expiring_soon"])
        self._api("get", "/api/dashboard", callback=cb)

    # ═══════════════════════════ KEYS ═══════════════════════════

    def _page_keys(self):
        p = QWidget()
        l = QVBoxLayout(p)
        l.setContentsMargins(24, 20, 24, 20)
        l.setSpacing(10)

        l.addWidget(QLabel("Лицензионные ключи", objectName="page-title"))

        # Generate form
        g = QGroupBox("Создать ключ")
        gl = QHBoxLayout(g)
        gl.setSpacing(6)
        self._k_org = QLineEdit(); self._k_org.setPlaceholderText("Организация*")
        self._k_phone = QLineEdit(); self._k_phone.setPlaceholderText("Телефон")
        self._k_city = QLineEdit(); self._k_city.setPlaceholderText("Город")
        self._k_cmt = QLineEdit(); self._k_cmt.setPlaceholderText("Комментарий")
        self._k_dur = QComboBox()
        self._k_dur.addItems(["7 дней", "14 дней", "1 месяц", "3 месяца", "6 месяцев", "12 месяцев"])
        self._k_dur.setCurrentIndex(2)
        for w in [self._k_org, self._k_phone, self._k_city, self._k_cmt, self._k_dur]:
            gl.addWidget(w)
        gen = QPushButton("Создать", objectName="primary")
        gen.clicked.connect(self._gen_key)
        gl.addWidget(gen)
        l.addWidget(g)

        # Toolbar
        tb = QHBoxLayout()
        tb.setSpacing(6)
        tb.addWidget(QLabel("Поиск:"))
        self._k_search = QLineEdit()
        self._k_search.setPlaceholderText("организация, телефон, ключ...")
        self._k_search.textChanged.connect(self._refresh_keys)
        tb.addWidget(self._k_search)
        tb.addWidget(QLabel("Статус:"))
        self._k_filter = QComboBox()
        self._k_filter.addItems(["Все", "active", "expired", "revoked"])
        self._k_filter.currentTextChanged.connect(self._refresh_keys)
        tb.addWidget(self._k_filter)
        tb.addStretch()
        bulk = QPushButton("Массовая генерация")
        bulk.clicked.connect(self._bulk_gen)
        tb.addWidget(bulk)
        exp = QPushButton("Excel")
        exp.clicked.connect(self._export_excel)
        tb.addWidget(exp)
        rev = QPushButton("Отозвать", objectName="warning")
        rev.clicked.connect(self._revoke_key)
        tb.addWidget(rev)
        dk = QPushButton("Удалить", objectName="danger")
        dk.clicked.connect(self._delete_key)
        tb.addWidget(dk)
        l.addLayout(tb)

        # Table
        self._k_table = QTableWidget()
        self._k_table.setColumnCount(9)
        self._k_table.setHorizontalHeaderLabels(["ID", "Ключ", "Организация", "Тел", "Город", "Комм", "Статус", "Создан", "Истекает"])
        self._k_table.setAlternatingRowColors(True)
        self._k_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._k_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._k_table.verticalHeader().setVisible(False)
        self._k_table.horizontalHeader().sectionClicked.connect(self._on_sort)
        self._k_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self._k_table.customContextMenuRequested.connect(self._key_menu)
        self._k_table.doubleClicked.connect(self._copy_key)
        widths = [35, 120, 140, 85, 65, 75, 90, 75, 75]
        for i, w in enumerate(widths):
            self._k_table.setColumnWidth(i, w)
        l.addWidget(self._k_table)
        return p

    def _populate_keys(self, keys):
        if self._sort_col >= 0:
            col_map = {0: "id", 1: "key", 2: "organization_name", 3: "phone",
                       4: "city", 5: "comment", 6: "status", 7: "created_at", 8: "expires_at"}
            sk = col_map.get(self._sort_col, "id")
            keys = sorted(keys, key=lambda k: str(k.get(sk, "")), reverse=not self._sort_asc)
        self._k_table.setRowCount(len(keys))
        st_map = {"active": "Активен", "expired": "Истёк", "revoked": "Отозван"}
        for i, k in enumerate(keys):
            st = st_map.get(k["status"], k["status"])
            if k.get("user_vk_id"):
                st += " (акт.)"
            vals = [str(k["id"]), k["key"], k["organization_name"],
                    k.get("phone", ""), k.get("city", ""), k.get("comment", ""),
                    st, (k.get("created_at", "") or "")[:10], (k.get("expires_at", "") or "")[:10]]
            for j, v in enumerate(vals):
                item = QTableWidgetItem(v)
                if j == 6:
                    item.setForeground(QColor({"Активен": "#3fb950", "Истёк": "#f85149", "Отозван": "#8b949e"}.get(st.split()[0], "#c9d1d9")))
                self._k_table.setItem(i, j, item)

    def _refresh_keys(self):
        st_map = {"Все": "", "active": "active", "expired": "expired", "revoked": "revoked"}
        st = st_map.get(self._k_filter.currentText(), "")
        search = self._k_search.text().strip()
        self._api("get", f"/api/keys?status={st}&search={search}", callback=lambda r: self._populate_keys(r.get("keys", [])))

    def _on_sort(self, col):
        if self._sort_col == col:
            self._sort_asc = not self._sort_asc
        else:
            self._sort_col = col
            self._sort_asc = True
        self._refresh_keys()

    def _key_menu(self, pos):
        row = self._k_table.currentRow()
        if row < 0:
            return
        m = QMenu(self)
        m.addAction("Копировать ключ", self._copy_key)
        m.addSeparator()
        m.addAction("Редактировать", self._edit_key)
        m.addAction("Отозвать", self._revoke_key)
        m.addAction("Удалить", self._delete_key)
        m.exec(self._k_table.viewport().mapToGlobal(pos))

    def _copy_key(self):
        row = self._k_table.currentRow()
        if row >= 0:
            key = self._k_table.item(row, 1).text()
            QApplication.clipboard().setText(key)
            self._status(f"Скопирован: {key}")

    def _edit_key(self):
        row = self._k_table.currentRow()
        if row < 0:
            return
        kid = int(self._k_table.item(row, 0).text())
        dlg = QDialog(self)
        dlg.setWindowTitle(f"Ключ #{kid}")
        dlg.setFixedSize(400, 340)
        dlg.setStyleSheet(STYLE)
        fl = QFormLayout(dlg)
        oe = QLineEdit(self._k_table.item(row, 2).text() or "")
        pe = QLineEdit((self._k_table.item(row, 3).text() if self._k_table.item(row, 3) else "") or "")
        ce = QLineEdit((self._k_table.item(row, 4).text() if self._k_table.item(row, 4) else "") or "")
        me = QLineEdit((self._k_table.item(row, 5).text() if self._k_table.item(row, 5) else "") or "")
        ex = QSpinBox(); ex.setRange(0, 365); ex.setValue(0); ex.setSuffix(" дн. (0=без изменений)")
        fl.addRow("Организация:", oe)
        fl.addRow("Телефон:", pe)
        fl.addRow("Город:", ce)
        fl.addRow("Комментарий:", me)
        fl.addRow("Продлить на:", ex)
        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        fl.addWidget(btns)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            data = {"id": kid, "organization_name": oe.text().strip(), "phone": pe.text().strip(),
                    "city": ce.text().strip(), "comment": me.text().strip()}
            if ex.value() > 0:
                data["expiry_days"] = ex.value()
            self._api("put", f"/api/keys/{kid}", data, callback=lambda r: (self._refresh_keys(), self._status("Обновлён")) if r.get("ok") else None)

    def _gen_key(self):
        org = self._k_org.text().strip()
        if not org:
            QMessageBox.warning(self, "Ошибка", "Введите организацию")
            return
        dur = {"7 дней": 7, "14 дней": 14, "1 месяц": 30, "3 месяца": 90, "6 месяцев": 180, "12 месяцев": 365}
        data = {"organization_name": org, "expiry_days": dur[self._k_dur.currentText()],
                "phone": self._k_phone.text().strip(), "city": self._k_city.text().strip(),
                "comment": self._k_cmt.text().strip()}
        def cb(r):
            if r.get("key"):
                QApplication.clipboard().setText(r["key"])
                self._k_org.clear(); self._k_phone.clear(); self._k_city.clear(); self._k_cmt.clear()
                self._refresh_keys()
                QMessageBox.information(self, "Ключ создан", f"Ключ: {r['key']}\n\nСкопирован в буфер.")
                self._status(f"Создан: {r['key']}")
            else:
                QMessageBox.critical(self, "Ошибка", "Не удалось создать")
        self._api("post", "/api/generate-key", data, callback=cb)

    def _bulk_gen(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Массовая генерация")
        dlg.setFixedSize(550, 350)
        dlg.setStyleSheet(STYLE)
        l = QVBoxLayout(dlg)
        l.addWidget(QLabel("CSV: org,phone,city,comment,days\n(org — обязательно, остальное опционально)"))
        te = QTextEdit()
        te.setPlaceholderText("ООО Ромашка,79001234567,Москва,Тест,30\nИП Иванов,,СПб,,90")
        l.addWidget(te)
        ds = QSpinBox(); ds.setRange(1, 365); ds.setValue(30); ds.setPrefix("Срок по умолчанию: "); ds.setSuffix(" дн.")
        l.addWidget(ds)
        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        l.addWidget(btns)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            entries = []
            default_days = ds.value()
            for line in te.toPlainText().strip().split("\n"):
                parts = [p.strip() for p in line.split(",")]
                if not parts or not parts[0]:
                    continue
                e = {"organization_name": parts[0], "expiry_days": default_days}
                if len(parts) >= 2 and parts[1]:
                    e["phone"] = parts[1]
                if len(parts) >= 3:
                    e["city"] = parts[2]
                if len(parts) >= 4:
                    e["comment"] = parts[3]
                if len(parts) >= 5:
                    try:
                        e["expiry_days"] = int(parts[4])
                    except ValueError:
                        pass
                entries.append(e)
            if entries:
                def cb(r):
                    if r.get("keys"):
                        QApplication.clipboard().setText("\n".join(r["keys"]))
                        self._refresh_keys()
                        QMessageBox.information(self, "Готово", f"Создано {r['count']} ключей.\n\nСкопированы в буфер.")
                self._api("post", "/api/generate-keys-bulk", {"entries": entries}, callback=cb)
            else:
                QMessageBox.warning(self, "Ошибка", "Нет данных")

    def _revoke_key(self):
        row = self._k_table.currentRow()
        if row < 0:
            return
        kid = int(self._k_table.item(row, 0).text())
        org = self._k_table.item(row, 2).text()
        if QMessageBox.question(self, "?", f"Отозвать ключ #{kid} ({org})?", QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No) == QMessageBox.StandardButton.Yes:
            self._api("post", "/api/revoke-key", {"key_id": kid}, callback=lambda r: (self._refresh_keys(), self._status(f"Отозван #{kid}")) if r.get("ok") else None)

    def _delete_key(self):
        row = self._k_table.currentRow()
        if row < 0:
            return
        kid = int(self._k_table.item(row, 0).text())
        org = self._k_table.item(row, 2).text()
        if QMessageBox.question(self, "?", f"Удалить ключ #{kid} ({org}) безвозвратно?", QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No) == QMessageBox.StandardButton.Yes:
            self._api("delete", f"/api/keys/{kid}", callback=lambda r: (self._refresh_keys(), self._status(f"Удалён #{kid}")) if r.get("ok") else None)

    def _export_excel(self):
        if not HAS_OPENPYXL:
            QMessageBox.warning(self, "Ошибка", "Установите openpyxl: pip install openpyxl")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Экспорт", "keys_export.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        def cb(r):
            keys = r.get("keys", [])
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Ключи"
            headers = ["ID", "Ключ", "Организация", "Телефон", "Город", "Комм.", "Статус", "Создан", "Истекает"]
            for c, h in enumerate(headers, 1):
                ws.cell(row=1, column=c, value=h)
            for i, k in enumerate(keys, 2):
                ws.cell(row=i, column=1, value=k["id"])
                ws.cell(row=i, column=2, value=k["key"])
                ws.cell(row=i, column=3, value=k["organization_name"])
                ws.cell(row=i, column=4, value=k.get("phone", ""))
                ws.cell(row=i, column=5, value=k.get("city", ""))
                ws.cell(row=i, column=6, value=k.get("comment", ""))
                ws.cell(row=i, column=7, value=k["status"])
                ws.cell(row=i, column=8, value=(k.get("created_at", "") or "")[:10])
                ws.cell(row=i, column=9, value=(k.get("expires_at", "") or "")[:10])
            wb.save(path)
            self._status(f"Экспортировано: {path}")
        self._api("get", "/api/keys", callback=cb)

    # ═══════════════════════════ EMPLOYEES ═══════════════════════════

    def _page_employees(self):
        p = QWidget()
        l = QVBoxLayout(p)
        l.setContentsMargins(24, 20, 24, 20)
        l.setSpacing(10)

        l.addWidget(QLabel("Сотрудники", objectName="page-title"))

        g = QGroupBox("Добавить")
        gl = QHBoxLayout(g)
        gl.setSpacing(6)
        self._e_fio = QLineEdit(); self._e_fio.setPlaceholderText("ФИО*")
        self._e_email = QLineEdit(); self._e_email.setPlaceholderText("Email*")
        self._e_pass = QLineEdit(); self._e_pass.setPlaceholderText("Пароль*"); self._e_pass.setEchoMode(QLineEdit.EchoMode.Password)
        self._e_phone = QLineEdit(); self._e_phone.setPlaceholderText("Телефон")
        self._e_role = QComboBox(); self._e_role.addItems(["operator", "manager", "admin"])
        for w in [self._e_fio, self._e_email, self._e_pass, self._e_phone, self._e_role]:
            gl.addWidget(w)
        add = QPushButton("Добавить", objectName="primary")
        add.clicked.connect(self._add_emp)
        gl.addWidget(add)
        l.addWidget(g)

        tb = QHBoxLayout()
        tb.addStretch()
        rp = QPushButton("Сбросить пароль")
        rp.clicked.connect(self._reset_emp)
        tb.addWidget(rp)
        de = QPushButton("Удалить", objectName="danger")
        de.clicked.connect(self._del_emp)
        tb.addWidget(de)
        l.addLayout(tb)

        self._e_table = QTableWidget()
        self._e_table.setColumnCount(6)
        self._e_table.setHorizontalHeaderLabels(["ID", "Email", "ФИО", "Телефон", "Роль", "Статус"])
        self._e_table.setAlternatingRowColors(True)
        self._e_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._e_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._e_table.verticalHeader().setVisible(False)
        for i, w in enumerate([35, 170, 150, 100, 75, 65]):
            self._e_table.setColumnWidth(i, w)
        l.addWidget(self._e_table)
        return p

    def _add_emp(self):
        data = {"fio": self._e_fio.text().strip(), "email": self._e_email.text().strip(),
                "password": self._e_pass.text().strip(), "phone": self._e_phone.text().strip(),
                "role": self._e_role.currentText()}
        if not data["fio"] or not data["email"] or not data["password"]:
            QMessageBox.warning(self, "Ошибка", "Заполните ФИО, Email и Пароль")
            return
        def cb(r):
            if r.get("ok"):
                self._e_fio.clear(); self._e_email.clear(); self._e_pass.clear(); self._e_phone.clear()
                self._refresh_employees()
                self._status("Добавлен")
            else:
                QMessageBox.critical(self, "Ошибка", r.get("error", "Email занят"))
        self._api("post", "/api/employees", data, callback=cb)

    def _refresh_employees(self):
        def cb(r):
            emps = r.get("employees", [])
            self._e_table.setRowCount(len(emps))
            for i, e in enumerate(emps):
                for j, k in enumerate(["id", "email", "fio", "phone", "role", "status"]):
                    item = QTableWidgetItem(str(e.get(k, "")))
                    if k == "status":
                        item.setForeground(QColor("#3fb950" if e[k] == "active" else "#f85149"))
                    self._e_table.setItem(i, j, item)
        self._api("get", "/api/employees", callback=cb)

    def _del_emp(self):
        row = self._e_table.currentRow()
        if row < 0:
            return
        eid = int(self._e_table.item(row, 0).text())
        name = self._e_table.item(row, 2).text()
        if QMessageBox.question(self, "?", f"Удалить {name}?", QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No) == QMessageBox.StandardButton.Yes:
            self._api("delete", f"/api/employees/{eid}", callback=lambda r: self._refresh_employees() if r.get("ok") else None)

    def _reset_emp(self):
        row = self._e_table.currentRow()
        if row < 0:
            return
        eid = int(self._e_table.item(row, 0).text())
        name = self._e_table.item(row, 2).text()
        np = f"reset{datetime.now().strftime('%H%M')}"
        self._api("post", "/api/auth/reset-password", {"emp_id": eid, "password": np},
                  callback=lambda r: QMessageBox.information(self, "Сброшен", f"{name}\nНовый пароль: {np}") if r.get("ok") else None)

    # ═══════════════════════════ AUDIT ═══════════════════════════

    def _page_audit(self):
        p = QWidget()
        l = QVBoxLayout(p)
        l.setContentsMargins(24, 20, 24, 20)
        l.setSpacing(10)

        l.addWidget(QLabel("Журнал аудита", objectName="page-title"))

        fl = QHBoxLayout()
        fl.setSpacing(6)
        fl.addWidget(QLabel("Действие:"))
        self._a_act = QComboBox()
        self._a_act.addItems(["Все", "login", "generate_key", "generate_keys_bulk", "revoke_key",
                               "delete_key", "update_key", "create_employee", "delete_employee",
                               "reset_password", "unbind_key"])
        self._a_act.currentTextChanged.connect(self._refresh_audit)
        fl.addWidget(self._a_act)
        fl.addWidget(QLabel("С:"))
        self._a_from = QDateEdit(); self._a_from.setDate(QDate.currentDate().addMonths(-1)); self._a_from.setCalendarPopup(True)
        fl.addWidget(self._a_from)
        fl.addWidget(QLabel("По:"))
        self._a_to = QDateEdit(); self._a_to.setDate(QDate.currentDate()); self._a_to.setCalendarPopup(True)
        fl.addWidget(self._a_to)
        r_btn = QPushButton("Обновить")
        r_btn.clicked.connect(self._refresh_audit)
        fl.addWidget(r_btn)
        fl.addStretch()
        l.addLayout(fl)

        self._a_table = QTableWidget()
        self._a_table.setColumnCount(6)
        self._a_table.setHorizontalHeaderLabels(["Время", "Пользователь", "Роль", "Действие", "Объект", "Детали"])
        self._a_table.setAlternatingRowColors(True)
        self._a_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._a_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._a_table.verticalHeader().setVisible(False)
        for i, w in enumerate([150, 120, 55, 130, 90, 230]):
            self._a_table.setColumnWidth(i, w)
        l.addWidget(self._a_table)
        return p

    def _refresh_audit(self):
        act = self._a_act.currentText()
        params = {
            "limit": 500,
            "date_from": self._a_from.date().toString("yyyy-MM-dd") + "T00:00:00",
            "date_to": self._a_to.date().toString("yyyy-MM-dd") + "T23:59:59",
        }
        if act != "Все":
            params["action"] = act
        def cb(r):
            log = r.get("log", [])
            self._a_table.setRowCount(len(log))
            for i, e in enumerate(log):
                ts = (e.get("created_at", "") or "")[:19].replace("T", " ")
                vals = [ts, e.get("user_name", "") or "", e.get("user_role", "") or "",
                        e.get("action", ""), f"{e.get('entity_type', '') or ''}#{e.get('entity_id', '') or ''}".rstrip("#"),
                        (e.get("details", "") or "")[:100]]
                for j, v in enumerate(vals):
                    self._a_table.setItem(i, j, QTableWidgetItem(v))
        self._api("get", "/api/audit", params, callback=cb)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    w = QyouroAdmin()
    sys.exit(app.exec())
